import openpyxl
import os
import shutil
from openpyxl.styles import PatternFill

def find_photos(code, folder):
    found_photos = []
    for root, dirs, files in os.walk(folder):
        for file in files:
            if file.lower().endswith('.jpg') and code in file:
                found_photos.append(os.path.join(root, file))
                if file.lower().startswith('(internal SKU code)') and code in file:
                    found_photos.append(os.path.join(root, file))
    return found_photos

def copy_photos_from_excel(excel_path, source_folder, destination_folder):
    if not os.path.exists(excel_path):
        print(f"Excel file '{excel_path}' not found.")
        return

    excel_file = openpyxl.load_workbook(excel_path)
    sheet = excel_file.active
    red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')

    if not os.path.exists(destination_folder):
        os.makedirs(destination_folder)

    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=2):
        code = row[0].value
        if code is None:
            print("Empty cell encountered. Skipping...")
            continue

        if any(code in file for file in os.listdir(destination_folder) if file.lower().endswith('.jpg')):
            print(f"Photo for code {code} already exists in the destination folder. Skipping...")
            continue

        source_paths = find_photos(code, source_folder)

        if source_paths:
            for source_path in source_paths:
                try:
                    shutil.copy(source_path, destination_folder)
                    print(f"Photo {os.path.basename(source_path)} copied successfully.")
                except Exception as e:
                    print(f"Failed to copy {os.path.basename(source_path)}: {e}")
        else:
            print(f"No photos found for code {code}.")
            row[0].fill = red_fill

    try:
        excel_file.save(excel_path)
        print(f"Excel file '{excel_path}' saved successfully.")
    except Exception as e:
        print(f"Failed to save Excel file '{excel_path}': {e}")
